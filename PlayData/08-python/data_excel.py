import openpyxl
filename = "stat_100701.xlsx"
book = openpyxl.load_workbook(filename)
sheet = book.active

data = []
for row in sheet.rows :
	data.append([
		row[0].value, 
		row[9].value
	])	# 2017년 시도별 인구수

for i in range(0,5) :		# 지우면 다음 인덱스가 0이 된다.
	del data[0]

# print( data )
data = sorted(data, key = lambda x:int( x[1].replace(",", "") ) , reverse=True)	# True = 내림차순
for i, d in enumerate(data) :
	if i >= 5 :
		break
	print( i+1, d[0], " : ", d[1])

savename = "population.xlsx"
wb = openpyxl.Workbook()
wbsheet = wb.active
wbsheet.title = "인구"

for i, d in enumerate(data) :
	wbsheet.cell(row = i+1, column=1, value=d[0])
	wbsheet.cell(row = i+1, column=2, value=d[1])
wb.save